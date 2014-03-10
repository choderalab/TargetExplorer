import sys, re
from openpyxl import Workbook
import pandas as pd

# ===========
# parameters
# ===========

PDB_cnstrcts_filename = 'PDB_constructs-data.csv'
output_Excel_filename = 'PDB_constructs.xlsx'


# ===========
# read in PDB construct data
# ===========

targets_results = pd.DataFrame.from_csv(PDB_cnstrcts_filename)

#print targets_results.keys
#print ''

cols_to_extract_from_targets_results = ['targetID', 'target_NCBI_GeneID', 'top_cnstrct_expr_tag', 'construct_aa_start', 'construct_aa_end', 'construct_aa_seq', 'construct_dna_start', 'construct_dna_end', 'construct_dna_seq']

# filter out targets for which the top-ranked PDB construct has an authenticity score of 0
targets_selection = targets_results[ targets_results['top_cnstrct_auth_score'] > 0 ]

# filter out targets for which the top-ranked plasmid has nconflicts > 0.
targets_selection = targets_selection[ targets_selection['top_plasmid_nconflicts'] == 0 ]

targets_selection = targets_selection[cols_to_extract_from_targets_results]

targets_selection.columns = ['targetID', 'GeneID', 'expression tag location', 'aa_start', 'aa_end', 'aa_seq', 'dna_start', 'dna_end', 'dna_seq']

# process the expression tag strings so they only say 'Nterm' or 'Cterm'
targets_selection['expression tag location'] = [ re.search('[CN]term', x).group() for x in targets_selection['expression tag location'] ]

# add 1 to each of the aa and dna start/end locations, so the values are 1-based
for col in ['aa_start', 'aa_end', 'dna_start', 'dna_end']:
    targets_selection[col] = [x+1 for x in targets_selection[col]]

#print targets_selection

targets_selection.to_excel(output_Excel_filename, index=False)

sys.exit()

# ===========
# write .xlsx spreadsheet containing the top 96 targets and necessary details for expression testing
# ===========

wb = Workbook()
ws = wb.get_active_sheet()
ws.title = output_Excel_filename[0 : output_Excel_filename.find('.xlsx')]

headings = ['targetID', 'GeneID', 'expression tag location', 'aa_start', 'aa_end', 'aa_seq', 'dna_start', 'dna_end', 'dna_seq']
for h in range(len(headings)):
    heading_cell = ws.cell(row=0, column=h)
    heading_cell.value = headings[h]

t_iter = 0
ntargets_selected = 0

# create new list of target results, removing targets with no matching PDB structures, and then keeping only the top [ndesired_unique_targets]
xl_output_targets_results = [t for t in targets_results if t['nmatching_PDB_structures'] > 0][0:ndesired_unique_targets]

# and a list of the same target results, sorted by DB rank
xl_output_targets_results_sorted_by_target_rank = sorted( xl_output_targets_results, key = lambda x: int(x['DB_target_rank']) )

for well_index in range(plate_size):
    if well_index < ndesired_unique_targets:
        target_index = well_index
        target_dict = xl_output_targets_results[target_index]
    else:
        target_index = well_index - ndesired_unique_targets
        target_dict = xl_output_targets_results_sorted_by_target_rank[target_index]

    targetID = target_dict['targetID']

    top_PDB_chain_ID = target_dict['top_PDB_chain_ID']

    target_NCBI_GeneID = target_dict['target_NCBI_GeneID']
    construct_aa_start = target_dict['construct_target_region_start_plasmid_coords']
    construct_aa_end = target_dict['construct_target_region_end_plasmid_coords'] # this refers to the last aa in 0-based aa coordinates
    construct_aa_seq = target_dict['construct_target_region_plasmid_seq']

    construct_dna_start = (construct_aa_start * 3)
    construct_dna_end = (construct_aa_end * 3) + 2 # this refers to the last nucleotide in 0-based nucleotide coordinates

    # Build spreadsheeet

    ID = targetID + '_' + top_PDB_chain_ID
    ID_cell = ws.cell(row=well_index+1, column=0)
    ID_cell.value = ID

    GeneID_cell = ws.cell(row=well_index+1, column=1)
    GeneID_cell.value = target_NCBI_GeneID
    
    exp_tag_loc = target_dict['top_construct_data']['tag_loc']
    exp_tag_loc_cell = ws.cell(row=well_index+1, column=2)
    exp_tag_loc_cell.value = exp_tag_loc

    # residue span: 1-based inclusive aa coordinates
    construct_aa_start_cell = ws.cell(row=well_index+1, column=3)
    construct_aa_end_cell = ws.cell(row=well_index+1, column=4)
    construct_aa_start_cell.value = construct_aa_start + 1
    construct_aa_end_cell.value = construct_aa_end + 1
    
    # aa_seq
    construct_aa_seq_cell = ws.cell(row=well_index+1, column=5)
    construct_aa_seq_cell.value = construct_aa_seq

    # DNA span: 1-based inclusive dna nucleotide coordinates
    construct_dna_start_cell = ws.cell(row=well_index+1, column=6)
    construct_dna_end_cell = ws.cell(row=well_index+1, column=7)
    construct_dna_start_cell.value = construct_dna_start + 1
    construct_dna_end_cell.value = construct_dna_end + 1
    
    # dna_seq
    orig_plasmid_dna_seq = plasmid_dna_seqs[target_NCBI_GeneID]
    construct_dna_seq = orig_plasmid_dna_seq[ construct_dna_start : construct_dna_end + 1 ]
    construct_dna_seq_cell = ws.cell(row=well_index+1, column=8)
    construct_dna_seq_cell.value = construct_dna_seq

    target_node = PDB_selections_xml_tree.find('target[@targetID="%s"]' % targetID)
    nreplicates = int(target_node.get('nreplicates')) + 1
    target_node.set('nreplicates', str(nreplicates))

    if len(construct_dna_seq) % 3 != 0:
        raise Exception, 'modulo 3 of DNA sequence length should be 0. Instead was %d' % (len(construct_dna_seq) % 3)

# Save spreadsheet
wb.save(output_Excel_filepath)


