# For a set of target protein domains, aligns PDB sequences against the target
# protein sequence and a sequence from a chosen plasmid library. Ranks the PDB
# sequences based on variables such as similarity to the target domain
# sequence, and likely authenticity of the PDB construct sequence (given that
# there are frequently errors in PDB SEQRES records). Outputs a list of chosen
# constructs in text format, a spreadsheet containing the constructs mapped
# onto the plasmid insert DNA sequences, and pretty HTML renderings of the
# alignments.
#
# Daniel L. Parton <partond@mskcc.org> - 3 Jan 2014
#

import sys, os, openpyxl, re, traceback
from lxml import etree
from lxml.builder import E
import TargetExplorer as clab
import Bio.Seq
import Bio.Alphabet
from openpyxl import Workbook
import pandas as pd

# ===========
# parameters
# ===========

try:
    override_target = sys.argv[ sys.argv.index('-target') + 1 ]
    targets = [override_target]
    plate_size = 1
    ndesired_unique_targets = 1
except ValueError:
    targets = 'All'
    # an Excel spreadsheet containing [plate_size] constructs will be created. The following value adjusts the number of unique targets. Any remaining well positions will be used for replicates, selected from the top-ranked DB entries (according to the DB target_score).
    plate_size = 96
    ndesired_unique_targets = 69

results_dir = os.path.join('analysis', 'PDB_construct_selection')

output_Excel_filename = 'PDB_constructs.xlsx'
output_selections_filename = 'PDB_constructs-data' # will be output as both .txt and .csv files
manual_exceptions_filename = 'manual_exceptions.yaml'
output_Excel_filepath = os.path.join(results_dir, output_Excel_filename)
output_selections_filepath = os.path.join(results_dir, output_selections_filename)
manual_exceptions_filepath = os.path.join(results_dir, manual_exceptions_filename)

html_alignments_dir = os.path.join(results_dir, 'alignments')

css_filename = 'seqlib.cs'
css_filepath = os.path.join(html_alignments_dir, css_filename)

output_columns = ['targetID', 'top_plasmid_nconflicts', 'nmatching_PDB_structures', 'top_PDB_chain_ID', 'top_cnstrct_expr_tag', 'top_cnstrct_auth_score', 'top_cnstrct_expr_sys', 'DB_target_rank', 'DB_target_score']

# ===========
# function definitions
# ===========

def generate_html_from_alignment(title, alignment, alignment_IDs, additional_data_fields=None, aa_css_class_list=None):
    '''
    additional_data_fields structure: [ { [PDB_ID]_[PDB_CHAIN_ID] : data }, { [PDB_ID]_[PDB_CHAIN_ID] : data }, ... ] with a separate dict for each data type, and where data is of len(alignment). Column headers not required.
    aa_css_class_list can be used to override the CSS classes assigned to each residue. Should be given as a list of lists, with shape: (len(alignment), len(alignment[0])).
    '''
    html_body = E.body()
    html_body.append( E.h2(title) )
    html_table = E.table()
    html_body.append( html_table )

    for i in range(len(alignment)):
        # row
        row = E.tr()

        # alignment ID div
        row.append( E.td( E.div(alignment_IDs[i],CLASS='ali') ) )

        # add any additional data fields (also in divs)
        for d in range(len(additional_data_fields)):
            data = additional_data_fields[d][alignment_IDs[i]]
            if data != None:
                row.append( E.td( E.div(data, CLASS='ali'), nowrap='') )
            else:
                row.append( E.td( E.div('', CLASS='ali'), nowrap='') )

        # format sequence with css classes. Returned as a list of span objects
        if aa_css_class_list != None:
            if aa_css_class_list[i] != None:
                prettyseq = clab.core.seq2pretty_html(alignment[i], aa_css_class_list=aa_css_class_list[i])
            else:
                prettyseq = clab.core.seq2pretty_html(alignment[i])
        else:
            prettyseq = clab.core.seq2pretty_html(alignment[i])

        # set up sequence div
        seq_div = E.div(id='sequence',CLASS='ali')
        seq_div.set('style','background-color:#dddddd;letter-spacing:-5px')

        # add sequence to div
        for span in prettyseq:
            seq_div.append(span)

        row.append( E.td( seq_div, nowrap='' ) )

        # add the row to the table
        html_table.append(row)

    return html_body

def match_regex(context, attrib_values, xpath_argument):
    # If no attrib found
    if len(attrib_values) == 0:
        return False
    # If attrib found, then run match against regex
    else:
        return bool( re.match(xpath_argument, attrib_values[0]) )

def write_output(ofilename, html_tree):
    ofile = open(ofilename, 'w')
    ofile.write( etree.tostring(html_tree, pretty_print=True) )
    ofile.close()

def process_target(t):
    target = targets_data[t].keys()[0]

    null_target_results = {
    'targetID' : target,
    'nmatching_PDB_structures' : 0,
    'top_PDB_chain_ID' : None,
    'top_cnstrct_expr_tag' : None,
    'top_cnstrct_auth_score' : 0,
    'top_cnstrct_expr_sys' : None,
    'target_NCBI_GeneID' : None,
    'construct_aa_start' : None,
    'construct_aa_end' : None,
    'construct_aa_seq' : None,
    'construct_dna_start' : None,
    'construct_dna_end' : None,
    'construct_dna_seq' : None,
    'top_plasmid_nconflicts' : None,
    'top_plasmid_pctidentity' : None,
    'target_domain_len' : None,
    'DB_target_score' : None,
    'DB_target_rank' : None
    }

    target_manual_exception = clab.core.parse_nested_dicts(manual_exceptions, [target, 'behavior'])
    if target_manual_exception != None:
        if target_manual_exception == 'skip':
            manual_exception_comment = clab.core.parse_nested_dicts(manual_exceptions, [target, 'comment'])
            print manual_exception_comment
            return null_target_results

    nmatching_PDB_structures = targets_data[t][target][0]
    print 'Working on target:', target
    # ===========
    # Generate html header
    # ===========
    output_html_tree = E.html(
        E.head(
            E.link()
        ),
        E.body(
        )
    )
    output_html_body = output_html_tree.find('body')
    css_link = output_html_tree.find('head/link')
    css_link.set('type','text/css')
    css_path = os.path.join(css_filename)
    css_link.set('href',css_path)
    css_link.set('rel','stylesheet')

    # ===========
    # Get target info from DB
    # ===========

    # get DB entry
    DB_domain = DB_root.find('entry/UniProt/domains/domain[@targetID="%s"]' % target)
    domainID = DB_domain.get('domainID')
    target_domain_len = int(DB_domain.get('length'))
    if target_domain_len > 350 or target_domain_len < 191:
        print 'Target domain length %d. Skipping...' % target_domain_len
        return null_target_results
    DB_entry = DB_domain.getparent().getparent().getparent()

    # get IDs
    target_UniProt_entry_name = DB_entry.find('UniProt').get('entry_name')
    target_NCBI_Gene_node = DB_entry.find('NCBI_Gene/entry[@ID]')
    if target_NCBI_Gene_node == None:
        print 'Gene ID not found for target %s' % target_UniProt_entry_name
        return null_target_results
    target_NCBI_GeneID = int(target_NCBI_Gene_node.get('ID'))

    #if target_NCBI_GeneID not in plasmid_data['NCBI_GeneID']:
    if sum(plasmid_data['NCBI_GeneID'] == target_NCBI_GeneID) == 0:
        print 'Gene ID %s (%s) not found in plasmid library.' % (target_NCBI_GeneID, target_UniProt_entry_name)
        return null_target_results

    # get UniProt canonical isoform sequence
    UniProt_canonseq = clab.core.sequnwrap( DB_entry.findtext('UniProt/isoforms/canonical_isoform/sequence') )

    # ===========
    # Get matching plasmids (via NCBI Gene ID)
    # ===========

    matching_plasmid_data = plasmid_data[ plasmid_data['NCBI_GeneID'] == target_NCBI_GeneID ]
    matching_plasmid_data.index = range(len(matching_plasmid_data))

    # ===========
    # Get PDB info from DB
    # ===========

    # get PDB sequences which correspond to the target domain and have the desired expression_system tag, and store in dict e.g. { '3GKZ_B' : 'MGYL...' }
    PDB_matching_seq_nodes = DB_entry.xpath( 'PDB/structure/expression_data[match_regex(@EXPRESSION_SYSTEM, "%s")]/../chain[@domainID="%s"]/experimental_sequence/sequence' % (desired_expression_system_regex, domainID), extensions = { (None, 'match_regex'): match_regex } )
    if len(PDB_matching_seq_nodes) == 0:
        return null_target_results
    # PDB_seqs structure: { [PDB_ID]_[PDB_CHAIN_ID] : sequence }
    # remove 'X' residues
    PDB_seqs = { seq_node.getparent().getparent().getparent().get('ID') + '_' + seq_node.getparent().getparent().get('ID') : clab.core.sequnwrap(seq_node.text.replace('X', '')) for seq_node in PDB_matching_seq_nodes }

    # PDB_seqs_aln_vs_UniProt_conflicts has same structure and contains the PDB construct seq "aligned" against the UniProt sequence, with conflicts in lower-case (derived using SIFTS info).
    # replace 'x' residues with '-'
    PDB_seqs_aln_vs_UniProt_conflicts = { seq_node.getparent().getparent().getparent().get('ID') + '_' + seq_node.getparent().getparent().get('ID') : clab.core.sequnwrap(seq_node.getparent().find('sequence_aln_conflicts').text.replace('x','-')) for seq_node in PDB_matching_seq_nodes }

    expression_system_data = { seq_node.getparent().getparent().getparent().get('ID') + '_' + seq_node.getparent().getparent().get('ID') : seq_node.getparent().getparent().getparent().find('expression_data').get('EXPRESSION_SYSTEM') for seq_node in PDB_matching_seq_nodes }

    # ===========
    # run MSA (using ClustalO)
    # ===========

    alignment_IDs = ['UniProt'] + list(matching_plasmid_data['cloneID'])
    #pre_alignment_seqs = [UniProt_canonseq, plasmid_aa_seqs[target_NCBI_GeneID]]
    pre_alignment_seqs = [UniProt_canonseq] + list(matching_plasmid_data['aa_seq'])

    for PDB_chain_ID in PDB_seqs.keys():
        alignment_IDs.append(PDB_chain_ID)
        pre_alignment_seqs.append(PDB_seqs[PDB_chain_ID])
    aligned_seqs = clab.align.run_clustalo(alignment_IDs, pre_alignment_seqs)
    alignment = [ [alignment_IDs[i], seq] for i, seq in enumerate(aligned_seqs) ] # convert aligned_seqs to this list of 2-element lists
    UniProt_aligned = alignment[0]
    alignment_plasmids_span = [1, 1 + len(matching_plasmid_data)]
    alignment_PDBs_span = [1 + len(matching_plasmid_data), None]

    # manual exceptions
    for i, PDB_chain_ID in enumerate(PDB_seqs.keys()):
        PDB_ID = PDB_chain_ID.split('_')[0]
        alignment_override = clab.core.parse_nested_dicts(manual_exceptions, [target, PDB_ID, 'alignment_override'])
        if alignment_override != None:
            alignment[alignment_PDBs_span[0] + i][1] = alignment_override.strip()

    # ===========
    # compare plasmid and PDB seqs with aligned UniProt canonical seq - put non-matching aas in lower case
    # ===========
    for seq_iter in range(len(alignment))[1:]:
        seq = list(alignment[seq_iter][1]) # sequence needs to be mutable so convert to list
        for aa_iter in range(len(seq)):
            if seq[aa_iter] != UniProt_aligned[1][aa_iter]:
                seq[aa_iter] = seq[aa_iter].lower()
        alignment[seq_iter][1] = ''.join(seq)

    # ===========
    # find the target domain within the aligned target domain seq
    # ===========
    target_domain_seq = clab.core.sequnwrap( DB_domain.findtext('sequence') )
    # to do this, we construct a regex which accounts for the possible presence of '-' chars within the target domain sequence.
    target_domain_seq_regex = ''.join([ aa + '-*' for aa in target_domain_seq ])[:-2] # ignore last '-*'
    aligned_UniProt_seq_target_domain_match = re.search(target_domain_seq_regex, UniProt_aligned[1])
    len_aligned_UniProt_seq_target_domain_match = aligned_UniProt_seq_target_domain_match.end() - aligned_UniProt_seq_target_domain_match.start()

    # ===========
    # set up data structures in preparation for sorting
    # ===========


    sorted_alignment = alignment[ 0 : alignment_plasmids_span[1] ] # the UniProt seq and plasmid seqs stay at the beginning
    plasmids_aligned = alignment[alignment_plasmids_span[0] : alignment_plasmids_span[1]]
    # PDB_construct_seqs_aligned structure: [ ['PDB_chain_ID', 'sequence'], ... ]
    PDB_construct_seqs_aligned = alignment[alignment_PDBs_span[0] : alignment_PDBs_span[1]]

    # ===========
    # now run pairwise alignments of the plasmid seqs against the UniProt canon seq target domain, calculate the number of conflicts, and sort the plasmid seqs based on nconflicts
    # ===========

    plasmids_nconflicts = []
    plasmids_pctidentity = []

    for p in matching_plasmid_data.index:

        pair_alignment_IDs = ['UniProt', matching_plasmid_data['cloneID'][p]]
        pair_pre_alignment_seqs = [target_domain_seq, matching_plasmid_data['aa_seq'][p]]
        pair_aligned_seqs = clab.align.run_clustalo(pair_alignment_IDs, pair_pre_alignment_seqs)
        pair_alignment = [ [pair_alignment_IDs[i], seq] for i, seq in enumerate(pair_aligned_seqs) ] # convert aligned_seqs to this list of 2-element lists

        regex_start = re.search('[A-Z]', pair_alignment[0][1])
        regex_end = re.search('[A-Z]', pair_alignment[0][1][::-1])
        target_domain_start_pair_aln_coords = regex_start.start()
        target_domain_end_pair_aln_coords = len(pair_alignment[0][1]) - regex_end.start() - 1  # 0-based index of the last residue

        pair_aln_seqs_target_domain_region = [pair_aln_seq[1][target_domain_start_pair_aln_coords : target_domain_end_pair_aln_coords + 1] for pair_aln_seq in pair_alignment]
        nconflicts = len( [ aa_iter for aa_iter in range(len(pair_aln_seqs_target_domain_region[0])) if pair_aln_seqs_target_domain_region[0][aa_iter] == '-' or pair_aln_seqs_target_domain_region[1][aa_iter] == '-' or (pair_aln_seqs_target_domain_region[0][aa_iter] != pair_aln_seqs_target_domain_region[1][aa_iter]) ] )
        pctidentity = (len(target_domain_seq) - nconflicts) * 100. / float(len(target_domain_seq))
        plasmids_nconflicts.append( nconflicts )
        plasmids_pctidentity.append( pctidentity )

    matching_plasmid_data['nconflicts'] = pd.Series(plasmids_nconflicts, index=matching_plasmid_data.index)
    matching_plasmid_data['pctidentity'] = pd.Series(plasmids_pctidentity, index=matching_plasmid_data.index)

    # sort plasmid seqs based on number of conflicts

    plasmids_aligned = sorted( plasmids_aligned, key = lambda x: matching_plasmid_data[ matching_plasmid_data['cloneID'] == x[0] ]['nconflicts'] )
    top_cloneID = plasmids_aligned[0][0]

    # put the sorted plasmids back into the main alignment lists
    alignment[alignment_plasmids_span[0] : alignment_plasmids_span[1]] = plasmids_aligned
    sorted_alignment[alignment_plasmids_span[0] : alignment_plasmids_span[1]] = plasmids_aligned

    # ===========
    # Detect extraneous expression tags - then calculate an "authenticity score", in an attempt to downrank misannotated construct sequences
    # ===========
    # Highest score for 'gh' within non-matching sequence (Abl1 E coli constructs use this at the N-term - comes after a His tag with TEV cleavage site)
    # Next search for His tag
    # Then look for non-matching sequence at the N- or C-term, of >3 aas.
    # Favor N-terminal tags, since QB3 MacroLab plasmids with His tags and TEV cleavage sites perform best in this configuration

    TEV_cleaved_Nterm_regex = '^g[has]m{0,1}g{0,1}[sd]{0,1}[A-Z]+[A-Z]{30}'
    TEV_uncleaved_Nterm_regex = '.*[eE][nNvV][lL][yY]{0,1}[fF][qQ].*[A-Z]{30}'
    TEV_Cterm_regex = '.*[A-Z]{30}.*[eE][nN][lL][yY][fF][qQ]'
    histag_Nterm_regex = '.*[hH]{6}.*[A-Z]+'
    histag_Cterm_regex = '.*[A-Z]+.*[hH]{6}'
    other_extra_seq_Nterm_regex = '.*[a-z]{3}.*[A-Z]{30}'
    other_extra_seq_Cterm_regex = '.*[A-Z]{30}.*[a-z]{3}'

    authenticity_scores = [0] * len(PDB_construct_seqs_aligned)
    # data for each construct will be added to this dict in the following for loop. Expression system data is added immediately.
    constructs_data = { x[0] : {'expression_system' : expression_system_data[x[0]]} for x in PDB_construct_seqs_aligned }
    expr_tag_strings = { x[0] : None for x in alignment }
    for i in range(len(PDB_construct_seqs_aligned)):
        ID = PDB_construct_seqs_aligned[i][0]
        PDB_entry_ID = ID.split('_')[0]
        seq = PDB_construct_seqs_aligned[i][1].replace('-', '') # remove '-' from sequence for regex searches

        # first check for manual exceptions
        manual_exception_behavior = clab.core.parse_nested_dicts(manual_exceptions, [target, PDB_entry_ID, 'authenticity_score', 'behavior'])

        override_tag_type = False
        if manual_exception_behavior != None:
            if manual_exception_behavior[0:8] == 'override':
                override_tag_type = manual_exception_behavior.split(';')[1].strip()

        if manual_exception_behavior == 'downweight':
            authenticity_scores[i] = -10
            expr_tag_strings[ID] = 'manually deprioritized'
            constructs_data[ID]['tag_type'] = None
            constructs_data[ID]['tag_loc'] = None
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue

        # now use regexes to check for the presence of expression tags, and use this information to set the authenticity_scores
        elif re.match(TEV_cleaved_Nterm_regex, seq) or override_tag_type == 'TEV_cleaved_Nterm':
            authenticity_scores[i] = 10
            expr_tag_strings[ID] = 'TEV_cleaved_Nterm'
            constructs_data[ID]['tag_type'] = 'TEV_cleaved'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(TEV_uncleaved_Nterm_regex, seq) or override_tag_type == 'TEV_uncleaved_Nterm':
            authenticity_scores[i] = 9
            expr_tag_strings[ID] = 'TEV_uncleaved_Nterm'
            constructs_data[ID]['tag_type'] = 'TEV_uncleaved'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(TEV_Cterm_regex, seq) or override_tag_type == 'TEV_Cterm':
            authenticity_scores[i] = 8
            expr_tag_strings[ID] = 'TEV_Cterm'
            constructs_data[ID]['tag_type'] = 'TEV'
            constructs_data[ID]['tag_loc'] = 'Cterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(histag_Nterm_regex, seq) or override_tag_type == 'Histag_Nterm':
            authenticity_scores[i] = 7
            expr_tag_strings[ID] = 'Histag_Nterm'
            constructs_data[ID]['tag_type'] = 'Histag'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(histag_Cterm_regex, seq) or override_tag_type == 'Histag_Cterm':
            authenticity_scores[i] = 6
            expr_tag_strings[ID] = 'Histag_Cterm'
            constructs_data[ID]['tag_type'] = 'Histag'
            constructs_data[ID]['tag_loc'] = 'Cterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(other_extra_seq_Nterm_regex, seq) or override_tag_type == 'other_extra_seq_Nterm':
            authenticity_scores[i] = 4
            expr_tag_strings[ID] = 'other_extra_seq_Nterm'
            constructs_data[ID]['tag_type'] = 'other_extra_seq'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(other_extra_seq_Cterm_regex, seq) or override_tag_type == 'other_extra_seq_Cterm':
            authenticity_scores[i] = 3
            expr_tag_strings[ID] = 'other_extra_seq_Cterm'
            constructs_data[ID]['tag_type'] = 'other_extra_seq'
            constructs_data[ID]['tag_loc'] = 'Cterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        else:
            authenticity_scores[i] = 0
            expr_tag_strings[ID] = None
            constructs_data[ID]['tag_type'] = None
            constructs_data[ID]['tag_loc'] = None
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue

    html_additional_data = pd.DataFrame( { x : [None] * len(alignment) for x in ['expr_system', 'expr_tag_string'] }, index=[a[0] for a in alignment] )

    for key in expression_system_data.keys():
        html_additional_data['expr_system'][key] = expression_system_data[key]
    for key in expr_tag_strings:
        html_additional_data['expr_tag_string'][key] = expr_tag_strings[key]

    # ===========
    # calculate the number of aas outside the target domain sequence
    # ===========

    num_aas_outside_target_domain = [0] * len(PDB_construct_seqs_aligned)
    for i in range(len(PDB_construct_seqs_aligned)):
        PDB_chain_ID = PDB_construct_seqs_aligned[i][0]
        # to get the total number of aas in the PDB construct, we take the mapping of that sequence against the UniProt sequence (derived from SIFTS data, not an alignment)
        # get this from the PDB construct sequence mapped (using SIFTS data - not an alignment) against the UniProt seq coords
        PDB_construct_start_UniProt_coords = re.search('[A-Za-z]', PDB_seqs_aln_vs_UniProt_conflicts[PDB_chain_ID]).start()
        PDB_construct_end_UniProt_coords = len(PDB_seqs_aln_vs_UniProt_conflicts[PDB_chain_ID]) - re.search('[A-Z]', PDB_seqs_aln_vs_UniProt_conflicts[PDB_chain_ID][::-1]).start()
        # this calculation will include '-' and 'x' chars within the span
        PDB_construct_len = PDB_construct_end_UniProt_coords - PDB_construct_start_UniProt_coords + 1
        num_aas_outside_target_domain[i] = PDB_construct_len - target_domain_len

    # ===========
    # score the alignment quality for the target domain sequence
    # ===========
    aln_scores = [0] * len(PDB_construct_seqs_aligned)
    for i in range(len(PDB_construct_seqs_aligned)):
        # extract target domain sequences
        UniProt_canon_seq_target_domain_seq = UniProt_aligned[1][ aligned_UniProt_seq_target_domain_match.start() : aligned_UniProt_seq_target_domain_match.end() ]
        PDB_seq_target_domain_seq = PDB_construct_seqs_aligned[i][1][ aligned_UniProt_seq_target_domain_match.start() : aligned_UniProt_seq_target_domain_match.end() ]
        # compare using PAM matrix
        aln_scores[i] = 0 - clab.align.score_aln(UniProt_canon_seq_target_domain_seq, PDB_seq_target_domain_seq) # subtract from 0 to reverse ordering

    # ===========
    # sort the PDB constructs based firstly on the authenticity_score (construct authenticity likelihood), then the alignment score, and finally on the number of aas outside the target domain sequence
    # ===========
    # dict_for_sorting structure: { sequence : (authenticity_score, alignment_score, num_aas_outside_target_domain), ... }
    # negate values for reverse sorting
    dict_for_sorting = { x[0] : (-authenticity_scores[i], aln_scores[i], num_aas_outside_target_domain[i]) for i, x in enumerate(PDB_construct_seqs_aligned) }
    # PDB_construct_seqs_aligned structure: [ ['PDB_chain_ID', 'sequence'], ... ]
    PDB_construct_seqs_aligned = sorted( PDB_construct_seqs_aligned, key = lambda x: dict_for_sorting[x[0]])
    sorted_alignment += PDB_construct_seqs_aligned
    top_PDB_chain_ID = PDB_construct_seqs_aligned[0][0]

    # ===========
    # generate HTML version of the alignment
    # ===========

    # generate custom set of css classes which will be used to highlight target domain of UniProt sequence in red ('c4')
    aa_css_class_list_UniProt_seq = [None] * len(alignment)
    aa_css_class_list_UniProt_seq[0] = ['bl'] * len(UniProt_aligned[1])
    aa_css_class_list_UniProt_seq[0][ aligned_UniProt_seq_target_domain_match.start() : aligned_UniProt_seq_target_domain_match.end() ] = ['c4'] * len_aligned_UniProt_seq_target_domain_match

    # generate the html
    sorted_alignment_seqs = [ x[1] for x in sorted_alignment ]
    sorted_alignment_IDs = [ x[0] for x in sorted_alignment ]
    # convert html_additional_data DataFrame back to list/dict format
    html_additional_data = [ { datatuple[0] : datatuple[1] for datatuple in series[1].iteritems() } for series in html_additional_data.iteritems() ]
    alignment_html = generate_html_from_alignment(target, sorted_alignment_seqs, sorted_alignment_IDs, additional_data_fields=html_additional_data, aa_css_class_list=aa_css_class_list_UniProt_seq)

    # add the alignment html to the main html tree
    for child in alignment_html.getchildren():
        output_html_body.append(child)

    # ===========
    # write to html file
    # ===========
    ofilename = os.path.join(html_alignments_dir, target + '.html')
    write_output(ofilename, output_html_tree)

    # ===========
    # find the "construct target region" (the central portion of the construct with sequence matching the target protein sequence, i.e. excluding expression tags etc.)
    # ===========

    # get this from the PDB construct sequence mapped (using SIFTS data - not an alignment) against the UniProt seq coords
    # desired span is from the first upper-case character to the last

    top_PDB_seq_aln_vs_UniProt_conflicts = PDB_seqs_aln_vs_UniProt_conflicts[top_PDB_chain_ID]
    regex_match_forward = re.search('[A-Z]', top_PDB_seq_aln_vs_UniProt_conflicts)
    top_PDB_seq_aln_vs_UniProt_conflicts_reverse = top_PDB_seq_aln_vs_UniProt_conflicts[::-1]
    regex_match_backward = re.search('[A-Z]', top_PDB_seq_aln_vs_UniProt_conflicts_reverse)
    construct_target_region_start_UniProt_coords = regex_match_forward.start()
    construct_target_region_end_UniProt_coords = len(top_PDB_seq_aln_vs_UniProt_conflicts) - regex_match_backward.end()
    construct_target_region_seq = top_PDB_seq_aln_vs_UniProt_conflicts[ construct_target_region_start_UniProt_coords : construct_target_region_end_UniProt_coords + 1]

    # now get the construct target region span in the coordinates of the alignment
    # do this by constructing a regex which accounts for the presence of '-' chars, and searching it against the aligned construct sequence
    # ignore '-' chars existing within construct_target_region_seq (which shouldn't be there according to the PDB standard, but frequently are, as SEQRES records often contain the observed sequence rather than the experimental sequence)
    # also convert all residues to upper case. This helps to avoid errors occurring due to non-ideal alignments.

    construct_target_region_regex = ''.join([ aa + '-*' for aa in construct_target_region_seq if aa != '-' ])[:-2] # ignore last '-*'
    regex_match = re.search(construct_target_region_regex.upper(), PDB_construct_seqs_aligned[0][1].upper())
    try:
        construct_target_region_start_aln_coords = regex_match.start()
    except Exception as e:
        print UniProt_canon_seq_target_domain_seq
        print construct_target_region_seq
        print construct_target_region_regex
        print PDB_construct_seqs_aligned[0][1]
        print PDB_construct_seqs_aligned[0][1].replace('-', '')
        print traceback.format_exc()
        raise e
    construct_target_region_end_aln_coords = regex_match.end() - 1

    # now get the top aligned plasmid sequence for the same span
    # TODO
    construct_target_region_aln_plasmid_seq = plasmids_aligned[0][1][ construct_target_region_start_aln_coords : construct_target_region_end_aln_coords + 1 ]
    # now get the (non-aligned) plasmid sequence; just need to remove '-' chars and convert conflicts back to upper case
    construct_target_region_plasmid_seq = construct_target_region_aln_plasmid_seq.replace('-', '').upper()
    # now get the start and end for the construct target region in the coords of the original plasmid seq
    top_plasmid_data = matching_plasmid_data[ matching_plasmid_data['cloneID'] == top_cloneID ]
    top_plasmid_data.index = [0]
    regex_match = re.search(construct_target_region_plasmid_seq, top_plasmid_data.aa_seq[0])
    construct_aa_start = regex_match.start()
    construct_aa_end = regex_match.end() - 1 # refers to the last residue in 0-based aa coords

    # ===========
    # get the target_score and target_rank from the DB
    # ===========

    DB_target_score_node = DB_root.find('entry/target_score/domain[@targetID="%s"]' % target)
    DB_target_score = DB_target_score_node.get('target_score')
    DB_target_rank = DB_target_score_node.get('target_rank')

    # ===========
    # extract construct span from plasmid DNA sequence
    # ===========

    construct_dna_start = (construct_aa_start * 3) # 0-based
    construct_dna_end = (construct_aa_end * 3) + 2 # this refers to the last nucleotide in 0-based nucleotide coordinates
    top_plasmid_dna_seq = top_plasmid_data.dna_seq[0]
    construct_dna_seq = top_plasmid_dna_seq[ construct_dna_start : construct_dna_end + 1 ]

    # ===========
    # add data to targets_results
    # ===========

    # Construct target results dict
    target_results = {}
    target_results['targetID'] = target
    target_results['nmatching_PDB_structures'] = nmatching_PDB_structures
    target_results['top_PDB_chain_ID'] = top_PDB_chain_ID
    # And the construct_data for the top_PDB_chain_ID
    top_construct_data = constructs_data[top_PDB_chain_ID]
    if top_construct_data['tag_type'] == None:
        target_results['top_cnstrct_expr_tag'] = None
    else:
        target_results['top_cnstrct_expr_tag'] = top_construct_data['tag_loc'] + '_' + top_construct_data['tag_type']
    target_results['top_cnstrct_auth_score'] = top_construct_data['authenticity_score']
    target_results['top_cnstrct_expr_sys'] = top_construct_data['expression_system']
    # And the Gene ID
    target_results['target_NCBI_GeneID'] = target_NCBI_GeneID
    # Add the plasmid sequence corresponding to the construct target region, and the start and end coordinates
    target_results['construct_aa_start'] = construct_aa_start
    target_results['construct_aa_end'] = construct_aa_end
    target_results['construct_aa_seq'] = construct_target_region_plasmid_seq
    target_results['construct_dna_start'] = construct_dna_start
    target_results['construct_dna_end'] = construct_dna_end
    target_results['construct_dna_seq'] = construct_dna_seq
    # And the number of conflicts between plasmid and UniProt canon seq within the target domain region
    matching_plasmid_data['nconflicts'] = pd.Series(plasmids_nconflicts, index=matching_plasmid_data.index)
    target_results['top_plasmid_nconflicts'] = top_plasmid_data.nconflicts[0]
    target_results['top_plasmid_pctidentity'] = top_plasmid_data.pctidentity[0]
    target_results['target_domain_len'] = len(target_domain_seq)
    # And the DB target_score and target_rank
    target_results['DB_target_score'] = DB_target_score
    target_results['DB_target_rank'] = DB_target_rank

    return target_results







# ===========
# Main
# ===========

if __name__ == '__main__':

    # ===========
    # Set up directories
    # ===========

    analysis_dir = 'analysis'
    if not os.path.exists(results_dir):
        os.mkdir(results_dir)
    if not os.path.exists(html_alignments_dir):
        os.mkdir(html_alignments_dir)

    # ===========
    # Parse DB
    # ===========

    DB_path = os.path.join('database', 'database.xml')
    DB_root = etree.parse(DB_path).getroot()
    nentries = len(DB_root)

    # ===========
    # Target selection
    # ===========

    if targets == 'All':
        targets = [ domain_node.get('targetID') for domain_node in DB_root.findall('entry/UniProt/domains/domain[@targetID]') ]

    desired_expression_system_regex = 'E.* COLI'

    # ===========
    # Parse manual exceptions file
    # ===========
    import yaml
    manual_exceptions = yaml.load( open(manual_exceptions_filepath, 'r').read() )

    # ===========
    # Write CSS stylesheet file
    # ===========
    clab.core.write_css_stylesheet(css_filepath)

    # ===========
    # Sort targets based on number of PDB structures with matching expression systems
    # ===========

    # targets_data and targets_results structure:[ { targetID : data } , ... ] where data will be constructed as [nmatching_PDB_structures, top_PDB_chain_ID, top_construct_data ]
    targets_data = []

    for target in targets:
        # get DB entry
        DB_domain = DB_root.find('entry/UniProt/domains/domain[@targetID="%s"]' % target)
        DB_entry = DB_domain.getparent().getparent().getparent()

        # get PDB structures
        matching_PDB_structures = DB_entry.xpath( 'PDB/structure/expression_data[match_regex(@EXPRESSION_SYSTEM, "%s")]' % desired_expression_system_regex, extensions = { (None, 'match_regex'): match_regex } )
        targets_data.append( { target : [ len(matching_PDB_structures) ] } )

    # ===========
    # Get Harvard plasmid library sequences
    # ===========

    plasmid_library_dir = os.path.join('external-data', 'plasmids', 'DFHCC-PlasmID', 'HIP-human_kinase_collection-pJP1520')
    Harvard_plasmid_library_filepath = os.path.join(plasmid_library_dir, 'Mehle_Kinase_VS1_pJP1520_new_plates.xlsx')

    wb = openpyxl.load_workbook(Harvard_plasmid_library_filepath)
    sheet_ranges = wb.get_sheet_by_name(name = 'Kinase_VS1_pJP1520_new_plates')
    nrows = sheet_ranges.get_highest_row()
    plasmid_data = {
    'cloneID':[],
    'NCBI_GeneID':[],
    'dna_seq':[],
    'aa_seq':[],
    }
    #plasmid_aa_seqs = {}
    #plasmid_dna_seqs = {}
    for row in range(2,nrows):
        cloneID = sheet_ranges.cell('E%d' % row).value
        NCBI_GeneID = sheet_ranges.cell('H%d' % row).value    # returns int
        Symbol = sheet_ranges.cell('I%d' % row).value
        dna_seq = sheet_ranges.cell('L%d' % row).value 
        if len(dna_seq) % 3 != 0:
            print 'WARNING: length of DNA sequence not divisible by 3, for plasmid with Gene Symbol %s and NCBI Gene ID %s' % (Symbol, NCBI_GeneID)
        # Translate to DNA sequence (don't include stop codon)
        aa_seq = Bio.Seq.Seq(dna_seq, Bio.Alphabet.generic_dna).translate(to_stop=True)
        #plasmid_aa_seqs[NCBI_GeneID] = aa_seq
        #plasmid_dna_seqs[NCBI_GeneID] = dna_seq
        plasmid_data['cloneID'].append(cloneID)
        plasmid_data['NCBI_GeneID'].append(NCBI_GeneID)
        plasmid_data['dna_seq'].append(dna_seq)
        plasmid_data['aa_seq'].append(str(aa_seq))

    plasmid_data = pd.DataFrame(plasmid_data)

    #plasmid_NCBI_GeneIDs = plasmid_aa_seqs.keys()

    print ''

    # ===========
    # iterate through targets
    # ===========

    from multiprocessing import Pool
    pool = Pool()
    targets_results = pool.map(process_target, range(len(targets_data)))
    targets_results = pd.DataFrame(targets_results)

    # ===========
    # sort targets based firstly on the PDB construct authenticity score, and secondly on the number of PDB constructs with the desired expression system
    # ===========
    targets_results = targets_results.sort( columns=['top_cnstrct_auth_score', 'nmatching_PDB_structures'], ascending=[False, False] )
    targets_results.index = range(len(targets_results))

    # ===========
    # print and write human-readable text file and CSV file, containing sorted targets with details on PDB structure selection
    # ===========

    PDB_selections_text = targets_results.to_string(columns=output_columns)
    with open(output_selections_filepath + '.txt', 'w') as output_selections_file:
        output_selections_file.write(PDB_selections_text)

    targets_results.to_csv(output_selections_filepath + '.csv')

    # print some statistics
    ntargets_with_plasmid = len( targets_results )
    ntargets_one_or_more_PDB = len( targets_results[targets_results['nmatching_PDB_structures'] > 0] )
    ntargets_zero_or_more_PDB_expr_tag = len( targets_results['top_cnstrct_auth_score'] > 0 )
    ntargets_one_or_more_PDB_expr_tag = len( targets_results[ (targets_results['nmatching_PDB_structures'] > 0) & (targets_results['top_cnstrct_auth_score'] > 0) ] )

    print '\nTotal targets with plasmid: %d\n' % ntargets_with_plasmid
    print 'Total targets with plasmid and > 1 matching PDB structures: %d\n' % ntargets_one_or_more_PDB
    print 'Total targets with plasmid and > 0 matching PDB structures and a detected expr_tag: %d\n' % ntargets_zero_or_more_PDB_expr_tag
    print 'Total targets with plasmid and > 1 matching PDB structures and a detected expr_tag: %d\n' % ntargets_one_or_more_PDB_expr_tag

