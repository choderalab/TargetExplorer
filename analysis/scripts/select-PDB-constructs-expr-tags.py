# Same as select-PDB_constructs.py but selects only PDB construct sequences which include a 6His tag. This is aimed at avoiding improperly annotated sequences.
#
# Daniel L. Parton <partond@mskcc.org> - 3 Jan 2014
#

import sys, os, openpyxl, re, collections, traceback
from lxml import etree
from lxml.builder import E
import TargetExplorer as clab
import Bio.Seq
import Bio.Alphabet
from openpyxl import Workbook

# ===========
# parameters
# ===========

try:
    override_target = sys.argv[ sys.argv.index('-target') + 1 ]
    targets = [override_target]
    plate_size = 1
    ndesired_unique_targets = 1
except ValueError:
    targets = 'All'
    # an Excel spreadsheet containing [plate_size] constructs will be created. The following value adjusts the number of unique targets. Any remaining well positions will be used for replicates, selected from the top-ranked DB entries (according to the DB target_score).
    plate_size = 96
    ndesired_unique_targets = 74

results_dir = os.path.join('analysis', 'PDB_construct_selection')

output_Excel_filename = 'PDB_constructs.xlsx'
output_selections_filename = 'PDB_constructs-data.txt'
manual_exceptions_filename = 'manual_exceptions.yaml'
output_Excel_filepath = os.path.join(results_dir, output_Excel_filename)
output_selections_filepath = os.path.join(results_dir, output_selections_filename)
manual_exceptions_filepath = os.path.join(results_dir, manual_exceptions_filename)

html_alignments_dir = os.path.join(results_dir, 'alignments')

css_filename = 'seqlib.cs'
css_filepath = os.path.join(html_alignments_dir, css_filename)

# ===========
# function definitions
# ===========

def generate_html_from_alignment(title, alignment, alignment_IDs, additional_data_fields=None, aa_css_class_list=None):
    '''
    additional_data_fields structure: [ { [PDB_ID]_[PDB_CHAIN_ID] : data }, { [PDB_ID]_[PDB_CHAIN_ID] : data }, ... ] with a separate dict for each data type, and where data is of len(alignment)
    aa_css_class_list can be used to override the CSS classes assigned to each residue. Should be given as a list of lists, with shape: (len(alignment), len(alignment[0])).
    '''
    html_body = E.body()
    html_body.append( E.h2(title) )
    html_table = E.table()
    html_body.append( html_table )

    for i in range(len(alignment)):
        # row
        row = E.tr()

        # alignment ID div
        row.append( E.td( E.div(alignment_IDs[i],CLASS='ali') ) )

        # add any additional data fields (also in divs)
        for d in range(len(additional_data_fields)):
            data = additional_data_fields[d][alignment_IDs[i]]
            if data != None:
                row.append( E.td( E.div(data, CLASS='ali'), nowrap='') )
            else:
                row.append( E.td( E.div('', CLASS='ali'), nowrap='') )

        # format sequence with css classes. Returned as a list of span objects
        if aa_css_class_list != None:
            if aa_css_class_list[i] != None:
                prettyseq = clab.core.seq2pretty_html(alignment[i], aa_css_class_list=aa_css_class_list[i])
            else:
                prettyseq = clab.core.seq2pretty_html(alignment[i])
        else:
            prettyseq = clab.core.seq2pretty_html(alignment[i])

        # set up sequence div
        seq_div = E.div(id='sequence',CLASS='ali')
        seq_div.set('style','background-color:#dddddd;letter-spacing:-5px')

        # add sequence to div
        for span in prettyseq:
            seq_div.append(span)

        row.append( E.td( seq_div, nowrap='' ) )

        # add the row to the table
        html_table.append(row)

    return html_body

def match_regex(context, attrib_values, xpath_argument):
    # If no attrib found
    if len(attrib_values) == 0:
        return False
    # If attrib found, then run match against regex
    else:
        return bool( re.match(xpath_argument, attrib_values[0]) )

def write_output(ofilename, html_tree):
    ofile = open(ofilename, 'w')
    ofile.write( etree.tostring(html_tree, pretty_print=True) )
    ofile.close()

def process_target(t):
    target = targets_data[t].keys()[0]

    # top PDB construct expression data will be added to a dict of this format (the variable is set here so that targets which are to be skipped can have data returned with the required structure)
    null_construct_data = {'expression_system' : None, 'tag_type' : None, 'tag_loc' : None, 'authenticity_score' : 0}
    # results for each target will be returned in this dict
    null_target_results = {'targetID' : target, 'nmatching_PDB_structures' : 0, 'top_PDB_chain_ID' : None, 'top_construct_data' : null_construct_data, 'target_NCBI_GeneID' : None, 'construct_target_region_start_plasmid_coords' : None, 'construct_target_region_end_plasmid_coords' : None, 'construct_target_region_plasmid_seq' : None, 'DB_target_score' : None, 'DB_target_rank' : None}

    nmatching_PDB_structures = targets_data[t][target][0]
    print 'Working on target:', target
    # ===========
    # Generate html header
    # ===========
    output_html_tree = E.html(
        E.head(
            E.link()
        ),
        E.body(
        )
    )
    output_html_body = output_html_tree.find('body')
    css_link = output_html_tree.find('head/link')
    css_link.set('type','text/css')
    css_path = os.path.join(css_filename)
    css_link.set('href',css_path)
    css_link.set('rel','stylesheet')

    # ===========
    # Get target info from DB
    # ===========

    # get DB entry
    DB_domain = DB_root.find('entry/UniProt/domains/domain[@targetID="%s"]' % target)
    target_domain_len = int(DB_domain.get('length'))
    if target_domain_len > 350 or target_domain_len < 191:
        print 'Target domain length %d. Skipping...' % target_domain_len
        return null_target_results
    DB_entry = DB_domain.getparent().getparent().getparent()

    # get IDs
    target_UniProt_entry_name = DB_entry.find('UniProt').get('entry_name')
    target_NCBI_Gene_node = DB_entry.find('NCBI_Gene/entry[@ID]')
    if target_NCBI_Gene_node == None:
        print 'Gene ID not found for target %s' % target_UniProt_entry_name
        return null_target_results
    target_NCBI_GeneID = int(target_NCBI_Gene_node.get('ID'))

    if target_NCBI_GeneID not in plasmid_NCBI_GeneIDs:
        print 'Gene ID %s (%s) not found in Harvard plasmid library.' % (target_NCBI_GeneID, target_UniProt_entry_name)
        return null_target_results

    # get UniProt canonical isoform sequence
    UniProt_canonseq = clab.core.sequnwrap( DB_entry.findtext('UniProt/isoforms/canonical_isoform/sequence') )

    # ===========
    # Get PDB info from DB
    # ===========

    # get PDB sequences (only those which have the desired expression_system tag) and store in dict e.g. { '3GKZ_B' : 'MGYL...' }
    PDB_matching_seq_nodes = DB_entry.xpath( 'PDB/structure/expression_data[match_regex(@EXPRESSION_SYSTEM, "%s")]/../chain/experimental_sequence/sequence' % desired_expression_system_regex, extensions = { (None, 'match_regex'): match_regex } )
    if len(PDB_matching_seq_nodes) == 0:
        return null_target_results
    # PDB_seqs structure: { [PDB_ID]_[PDB_CHAIN_ID] : sequence }
    # remove 'X' residues
    PDB_seqs = { seq_node.getparent().getparent().getparent().get('ID') + '_' + seq_node.getparent().getparent().get('ID') : clab.core.sequnwrap(seq_node.text.replace('X', '')) for seq_node in PDB_matching_seq_nodes }

    # PDB_seqs_aln_vs_UniProt_conflicts has same structure and contains the PDB construct seq "aligned" against the UniProt sequence, with conflicts in lower-case (derived using SIFTS info).
    # replace 'x' residues with '-'
    PDB_seqs_aln_vs_UniProt_conflicts = { seq_node.getparent().getparent().getparent().get('ID') + '_' + seq_node.getparent().getparent().get('ID') : clab.core.sequnwrap(seq_node.getparent().find('sequence_aln_conflicts').text.replace('x','-')) for seq_node in PDB_matching_seq_nodes }

    # html_additional_data structure: [ { [PDB_ID]_[PDB_CHAIN_ID] : data }, { [PDB_ID]_[PDB_CHAIN_ID] : data }, ... ] with a separate dict for each data type, and where data is of len(alignment)
    expression_system_data = { seq_node.getparent().getparent().getparent().get('ID') + '_' + seq_node.getparent().getparent().get('ID') : seq_node.getparent().getparent().getparent().find('expression_data').get('EXPRESSION_SYSTEM') for seq_node in PDB_matching_seq_nodes }
    html_additional_data = [{'UniProt': '', 'plasmid_seq': ''}] # No expression_system data to show for UniProt and plasmid sequences
    for key in expression_system_data.keys():
        html_additional_data[0][key] = expression_system_data[key]

    # ===========
    # run MSA (using ClustalO)
    # ===========
    alignment_IDs = ['UniProt', 'plasmid_seq']
    pre_alignment_seqs = [UniProt_canonseq, plasmid_aa_seqs[target_NCBI_GeneID]]
    for PDB_chain_ID in PDB_seqs.keys():
        alignment_IDs.append(PDB_chain_ID)
        pre_alignment_seqs.append(PDB_seqs[PDB_chain_ID])
    aligned_seqs = clab.align.run_clustalo(alignment_IDs, pre_alignment_seqs)
    alignment = [ [alignment_IDs[i], seq] for i, seq in enumerate(aligned_seqs) ] # convert aligned_seqs to this list of 2-element lists
    UniProt_canon_seq_aligned = alignment[0]

    # manual exceptions
    for i, PDB_chain_ID in enumerate(PDB_seqs.keys()):
        PDB_ID = PDB_chain_ID.split('_')[0]
        alignment_override = clab.core.parse_nested_dicts(manual_exceptions, [target, PDB_ID, 'alignment_override'])
        if alignment_override != None:
            alignment[i+2][1] = alignment_override.strip()

    # ===========
    # compare plasmid and PDB seqs with aligned UniProt canonical seq - put non-matching aas in lower case
    # ===========
    for seq_iter in range(len(alignment))[1:]:
        seq = list(alignment[seq_iter][1]) # sequence needs to be mutable so convert to list
        for aa_iter in range(len(seq)):
            if seq[aa_iter] != UniProt_canon_seq_aligned[1][aa_iter]:
                seq[aa_iter] = seq[aa_iter].lower()
        alignment[seq_iter][1] = ''.join(seq)

    # ===========
    # find the target domain within the aligned target domain seq
    # ===========
    target_domain_seq = clab.core.sequnwrap( DB_domain.findtext('sequence') )
    # to do this, we construct a regex which accounts for the possible presence of '-' chars within the target domain sequence.
    target_domain_seq_regex = ''.join([ aa + '-*' for aa in target_domain_seq ])[:-2] # ignore last '-*'
    aligned_UniProt_seq_target_domain_match = re.search(target_domain_seq_regex, UniProt_canon_seq_aligned[1])
    len_aligned_UniProt_seq_target_domain_match = aligned_UniProt_seq_target_domain_match.end() - aligned_UniProt_seq_target_domain_match.start()

    # ===========
    # set up data structures in preparation for sorting
    # ===========
    sorted_alignment = alignment[0:2] # the UniProt seq and plasmid seq stay at the beginning
    # PDB_construct_seqs_aligned structure: [ ['PDB_chain_ID', 'sequence'], ... ]
    PDB_construct_seqs_aligned = alignment[2:]

    # ===========
    # Detect extraneous expression tags - then calculate an "authenticity score", in an attempt to downrank misannotated construct sequences
    # ===========
    # Highest score for 'gh' within non-matching sequence (Abl1 E coli constructs use this at the N-term - comes after a His tag with TEV cleavage site)
    # Next search for His tag
    # Then look for non-matching sequence at the N- or C-term, of >3 aas.
    # Favor N-terminal tags, since QB3 MacroLab plasmids with His tags and TEV cleavage sites perform best in this configuration

    TEV_cleaved_Nterm_regex = '^g[has]m{0,1}g{0,1}[sd]{0,1}[A-Z]+[A-Z]{30}'
    TEV_uncleaved_Nterm_regex = '.*[eE][nNvV][lL][yY]{0,1}[fF][qQ].*[A-Z]{30}'
    TEV_Cterm_regex = '.*[A-Z]{30}.*[eE][nN][lL][yY][fF][qQ]'
    histag_Nterm_regex = '.*[hH]{6}.*[A-Z]+'
    histag_Cterm_regex = '.*[A-Z]+.*[hH]{6}'
    other_extra_seq_Nterm_regex = '.*[a-z]{3}.*[A-Z]{30}'
    other_extra_seq_Cterm_regex = '.*[A-Z]{30}.*[a-z]{3}'

    authenticity_scores = [0] * len(PDB_construct_seqs_aligned)
    # data for each construct will be added to this dict in the following for loop. Expression system data is added immediately.
    constructs_data = { x[0] : {'expression_system' : expression_system_data[x[0]]} for x in alignment[2:] }
    expr_tag_strings = { x[0] : None for x in alignment }
    for i in range(len(PDB_construct_seqs_aligned)):
        ID = PDB_construct_seqs_aligned[i][0]
        PDB_entry_ID = ID.split('_')[0]
        seq = PDB_construct_seqs_aligned[i][1].replace('-', '') # remove '-' from sequence for regex searches

        # first check for manual exceptions
        manual_exception_behavior = clab.core.parse_nested_dicts(manual_exceptions, [target, PDB_entry_ID, 'authenticity_score', 'behavior'])

        override_tag_type = False
        if manual_exception_behavior != None:
            if manual_exception_behavior[0:8] == 'override':
                override_tag_type = manual_exception_behavior.split(';')[1].strip()

        if manual_exception_behavior == 'downweight':
            authenticity_scores[i] = -10
            expr_tag_strings[ID] = 'manually deprioritized'
            constructs_data[ID]['tag_type'] = None
            constructs_data[ID]['tag_loc'] = None
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue

        # now use regexes to check for the presence of expression tags, and use this information to set the authenticity_scores
        elif re.match(TEV_cleaved_Nterm_regex, seq) or override_tag_type == 'TEV_cleaved_Nterm':
            authenticity_scores[i] = 10
            expr_tag_strings[ID] = 'TEV_cleaved_Nterm'
            constructs_data[ID]['tag_type'] = 'TEV_cleaved'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(TEV_uncleaved_Nterm_regex, seq) or override_tag_type == 'TEV_uncleaved_Nterm':
            authenticity_scores[i] = 9
            expr_tag_strings[ID] = 'TEV_uncleaved_Nterm'
            constructs_data[ID]['tag_type'] = 'TEV_uncleaved'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(TEV_Cterm_regex, seq) or override_tag_type == 'TEV_Cterm':
            authenticity_scores[i] = 8
            expr_tag_strings[ID] = 'TEV_Cterm'
            constructs_data[ID]['tag_type'] = 'TEV'
            constructs_data[ID]['tag_loc'] = 'Cterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(histag_Nterm_regex, seq) or override_tag_type == 'Histag_Nterm':
            authenticity_scores[i] = 7
            expr_tag_strings[ID] = 'Histag_Nterm'
            constructs_data[ID]['tag_type'] = 'Histag'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(histag_Cterm_regex, seq) or override_tag_type == 'Histag_Cterm':
            authenticity_scores[i] = 6
            expr_tag_strings[ID] = 'Histag_Cterm'
            constructs_data[ID]['tag_type'] = 'Histag'
            constructs_data[ID]['tag_loc'] = 'Cterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(other_extra_seq_Nterm_regex, seq) or override_tag_type == 'other_extra_seq_Nterm':
            authenticity_scores[i] = 4
            expr_tag_strings[ID] = 'other_extra_seq_Nterm'
            constructs_data[ID]['tag_type'] = 'other_extra_seq'
            constructs_data[ID]['tag_loc'] = 'Nterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        elif re.match(other_extra_seq_Cterm_regex, seq) or override_tag_type == 'other_extra_seq_Cterm':
            authenticity_scores[i] = 3
            expr_tag_strings[ID] = 'other_extra_seq_Cterm'
            constructs_data[ID]['tag_type'] = 'other_extra_seq'
            constructs_data[ID]['tag_loc'] = 'Cterm'
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue
        else:
            authenticity_scores[i] = 0
            expr_tag_strings[ID] = None
            constructs_data[ID]['tag_type'] = None
            constructs_data[ID]['tag_loc'] = None
            constructs_data[ID]['authenticity_score'] = authenticity_scores[i]
            continue

    html_additional_data.append( expr_tag_strings )

    # ===========
    # calculate the number of aas outside the target domain sequence
    # ===========

    num_aas_outside_target_domain = [0] * len(PDB_construct_seqs_aligned)
    for i in range(len(PDB_construct_seqs_aligned)):
        PDB_chain_ID = PDB_construct_seqs_aligned[i][0]
        # to get the total number of aas in the PDB construct, we take the mapping of that sequence against the UniProt sequence (derived from SIFTS data, not an alignment)
        # get this from the PDB construct sequence mapped (using SIFTS data - not an alignment) against the UniProt seq coords
        PDB_construct_start_UniProt_coords = re.search('[A-Za-z]', PDB_seqs_aln_vs_UniProt_conflicts[PDB_chain_ID]).start()
        PDB_construct_end_UniProt_coords = len(PDB_seqs_aln_vs_UniProt_conflicts[PDB_chain_ID]) - re.search('[A-Z]', PDB_seqs_aln_vs_UniProt_conflicts[PDB_chain_ID][::-1]).start()
        # this calculation will include '-' and 'x' chars within the span
        PDB_construct_len = PDB_construct_end_UniProt_coords - PDB_construct_start_UniProt_coords + 1
        num_aas_outside_target_domain[i] = PDB_construct_len - target_domain_len

    # ===========
    # score the alignment quality for the target domain sequence
    # ===========
    aln_scores = [0] * len(PDB_construct_seqs_aligned)
    for i in range(len(PDB_construct_seqs_aligned)):
        # extract target domain sequences
        UniProt_canon_seq_target_domain_seq = UniProt_canon_seq_aligned[1][ aligned_UniProt_seq_target_domain_match.start() : aligned_UniProt_seq_target_domain_match.end() ]
        PDB_seq_target_domain_seq = PDB_construct_seqs_aligned[i][1][ aligned_UniProt_seq_target_domain_match.start() : aligned_UniProt_seq_target_domain_match.end() ]
        # compare using PAM matrix
        aln_scores[i] = 0 - clab.align.score_aln(UniProt_canon_seq_target_domain_seq, PDB_seq_target_domain_seq) # subtract from 0 to reverse ordering

    # ===========
    # sort the PDB constructs based firstly on the authenticity_score (construct authenticity likelihood), then the alignment score, and finally on the number of aas outside the target domain sequence
    # ===========
    # dict_for_sorting structure: { sequence : (authenticity_score, alignment_score, num_aas_outside_target_domain), ... }
    # negate values for reverse sorting
    dict_for_sorting = { x[0] : (-authenticity_scores[i], aln_scores[i], num_aas_outside_target_domain[i]) for i, x in enumerate(PDB_construct_seqs_aligned) }
    # PDB_construct_seqs_aligned structure: [ ['PDB_chain_ID', 'sequence'], ... ]
    PDB_construct_seqs_aligned = sorted( PDB_construct_seqs_aligned, key = lambda x: dict_for_sorting[x[0]])
    sorted_alignment += PDB_construct_seqs_aligned
    top_PDB_chain_ID = PDB_construct_seqs_aligned[0][0]

    # ===========
    # generate custom set of css classes which will be used to highlight target domain of UniProt sequence in red ('c4')
    # ===========
    aa_css_class_list_UniProt_seq = [None] * len(alignment)
    aa_css_class_list_UniProt_seq[0] = ['bl'] * len(UniProt_canon_seq_aligned[1])
    aa_css_class_list_UniProt_seq[0][ aligned_UniProt_seq_target_domain_match.start() : aligned_UniProt_seq_target_domain_match.end() ] = ['c4'] * len_aligned_UniProt_seq_target_domain_match

    # ===========
    # generate html for the alignment
    # ===========
    sorted_alignment_seqs = [ x[1] for x in sorted_alignment ]
    sorted_alignment_IDs = [ x[0] for x in sorted_alignment ]
    alignment_html = generate_html_from_alignment(target, sorted_alignment_seqs, sorted_alignment_IDs, additional_data_fields=html_additional_data, aa_css_class_list=aa_css_class_list_UniProt_seq)

    # add the alignment html to the main html tree
    for child in alignment_html.getchildren():
        output_html_body.append(child)

    # ===========
    # write to html file
    # ===========
    ofilename = os.path.join(html_alignments_dir, target + '.html')
    write_output(ofilename, output_html_tree)

    # ===========
    # find the "construct target region" (the central portion of the construct with sequence matching the target protein sequence, i.e. excluding expression tags etc.)
    # ===========

    # get this from the PDB construct sequence mapped (using SIFTS data - not an alignment) against the UniProt seq coords
    # desired span is from the first upper-case character to the last

    top_PDB_seq_aln_vs_UniProt_conflicts = PDB_seqs_aln_vs_UniProt_conflicts[top_PDB_chain_ID]
    regex_match_forward = re.search('[A-Z]', top_PDB_seq_aln_vs_UniProt_conflicts)
    top_PDB_seq_aln_vs_UniProt_conflicts_reverse = top_PDB_seq_aln_vs_UniProt_conflicts[::-1]
    regex_match_backward = re.search('[A-Z]', top_PDB_seq_aln_vs_UniProt_conflicts_reverse)
    construct_target_region_start_UniProt_coords = regex_match_forward.start()
    construct_target_region_end_UniProt_coords = len(top_PDB_seq_aln_vs_UniProt_conflicts) - regex_match_backward.end()
    construct_target_region_seq = top_PDB_seq_aln_vs_UniProt_conflicts[ construct_target_region_start_UniProt_coords : construct_target_region_end_UniProt_coords + 1]

    # now get the construct target region span in the coordinates of the alignment
    # do this by constructing a regex which accounts for the presence of '-' chars, and searching it against the aligned construct sequence
    # ignore '-' chars existing within construct_target_region_seq (which shouldn't be there according to the PDB standard, but frequently are, as SEQRES records often contain the observed sequence rather than the experimental sequence)
    # also convert all residues to upper case. This helps to avoid errors occurring due to non-ideal alignments.

    construct_target_region_regex = ''.join([ aa + '-*' for aa in construct_target_region_seq if aa != '-' ])[:-2] # ignore last '-*'
    regex_match = re.search(construct_target_region_regex.upper(), sorted_alignment_seqs[2].upper())
    try:
        construct_target_region_start_aln_coords = regex_match.start()
    except Exception as e:
        print UniProt_canon_seq_target_domain_seq
        print construct_target_region_seq
        print construct_target_region_regex
        print sorted_alignment_seqs[2]
        print sorted_alignment_seqs[2].replace('-', '')
        print traceback.format_exc()
        raise e
    construct_target_region_end_aln_coords = regex_match.end() - 1

    # now get the aligned plasmid sequence for the same span
    construct_target_region_aln_plasmid_seq = sorted_alignment_seqs[1][ construct_target_region_start_aln_coords : construct_target_region_end_aln_coords + 1 ]
    # now get the (non-aligned) plasmid sequence; just need to remove '-' chars and convert conflicts back to upper case
    construct_target_region_plasmid_seq = construct_target_region_aln_plasmid_seq.replace('-', '').upper()
    # now get the start and end for the construct target region in the coords of the original plasmid seq
    regex_match = re.search(construct_target_region_plasmid_seq, str(plasmid_aa_seqs[target_NCBI_GeneID]))
    construct_target_region_start_plasmid_coords = regex_match.start()
    construct_target_region_end_plasmid_coords = regex_match.end() - 1

    # ===========
    # get the target_score and target_rank from the DB
    # ===========

    DB_target_score_node = DB_root.find('entry/target_score/domain[@targetID="%s"]' % target)
    DB_target_score = DB_target_score_node.get('target_score')
    DB_target_rank = DB_target_score_node.get('target_rank')

    # ===========
    # add data to targets_results
    # ===========

    # Construct target results dict
    target_results = {}
    target_results['targetID'] = target
    target_results['nmatching_PDB_structures'] = nmatching_PDB_structures
    target_results['top_PDB_chain_ID'] = top_PDB_chain_ID
    # And the construct_data for the top_PDB_chain_ID
    top_construct_data = constructs_data[top_PDB_chain_ID]
    target_results['top_construct_data'] = top_construct_data
    # And the Gene ID
    target_results['target_NCBI_GeneID'] = target_NCBI_GeneID
    # Add the plasmid sequence corresponding to the construct target region, and the start and end aa coordinates
    target_results['construct_target_region_start_plasmid_coords'] = construct_target_region_start_plasmid_coords
    target_results['construct_target_region_end_plasmid_coords'] = construct_target_region_end_plasmid_coords
    target_results['construct_target_region_plasmid_seq'] = construct_target_region_plasmid_seq
    # And the DB target_score and target_rank
    target_results['DB_target_score'] = DB_target_score
    target_results['DB_target_rank'] = DB_target_rank

    return target_results







# ===========
# Main
# ===========

if __name__ == '__main__':

    # ===========
    # Set up directories
    # ===========

    analysis_dir = 'analysis'
    if not os.path.exists(results_dir):
        os.mkdir(results_dir)
    if not os.path.exists(html_alignments_dir):
        os.mkdir(html_alignments_dir)

    # ===========
    # Parse DB
    # ===========

    DB_path = os.path.join('database', 'database.xml')
    DB_root = etree.parse(DB_path).getroot()
    nentries = len(DB_root)

    # ===========
    # Target selection
    # ===========

    if targets == 'All':
        targets = [ domain_node.get('targetID') for domain_node in DB_root.findall('entry/UniProt/domains/domain[@targetID]') ]

    desired_expression_system_regex = 'E.* COLI'

    # ===========
    # Parse manual exceptions file
    # ===========
    import yaml
    manual_exceptions = yaml.load( open(manual_exceptions_filepath, 'r').read() )

    # ===========
    # Parse manual exceptions file
    # ===========
    clab.core.write_css_stylesheet(css_filepath)

    # ===========
    # Sort targets based on number of PDB structures with matching expression systems
    # ===========

    # targets_data and targets_results structure:[ { targetID : data } , ... ] where data will be constructed as [nmatching_PDB_structures, top_PDB_chain_ID, top_construct_data ]
    targets_data = []

    for target in targets:
        # get DB entry
        DB_domain = DB_root.find('entry/UniProt/domains/domain[@targetID="%s"]' % target)
        DB_entry = DB_domain.getparent().getparent().getparent()

        # get PDB structures
        matching_PDB_structures = DB_entry.xpath( 'PDB/structure/expression_data[match_regex(@EXPRESSION_SYSTEM, "%s")]' % desired_expression_system_regex, extensions = { (None, 'match_regex'): match_regex } )
        targets_data.append( { target : [ len(matching_PDB_structures) ] } )

    # ===========
    # Get Harvard plasmid library sequences
    # ===========

    plasmid_library_dir = os.path.join('external-data', 'Harvard-DFHCC-HIP_human_kinase_collection-pJP1520')
    Harvard_plasmid_library_filepath = os.path.join(plasmid_library_dir, 'Mehle_Kinase_VS1_pJP1520_new_plates.xlsx')

    wb = openpyxl.load_workbook(Harvard_plasmid_library_filepath)
    sheet_ranges = wb.get_sheet_by_name(name = 'Kinase_VS1_pJP1520_new_plates')
    nrows = sheet_ranges.get_highest_row()
    plasmid_aa_seqs = {}
    plasmid_dna_seqs = {}
    for row in range(2,nrows):
        NCBI_GeneID = sheet_ranges.cell('H%d' % row).value    # type(int)
        Symbol = sheet_ranges.cell('I%d' % row).value
        dna_seq = sheet_ranges.cell('L%d' % row).value 
        if len(dna_seq) % 3 != 0:
            print 'WARNING: length of DNA sequence not divisible by 3, for plasmid with Gene Symbol %s and NCBI Gene ID %s' % (Symbol, NCBI_GeneID)
        # Translate to DNA sequence (don't include stop codon)
        aa_seq = Bio.Seq.Seq(dna_seq, Bio.Alphabet.generic_dna).translate(to_stop=True)
        plasmid_aa_seqs[NCBI_GeneID] = aa_seq
        plasmid_dna_seqs[NCBI_GeneID] = dna_seq

    plasmid_NCBI_GeneIDs = plasmid_aa_seqs.keys()

    print ''

    # ===========
    # iterate through targets
    # ===========

    from multiprocessing import Pool
    pool = Pool()
    targets_results = pool.map(process_target, range(len(targets_data)))

    # ===========
    # sort targets based firstly on the PDB construct authenticity score, and secondly on the number of PDB constructs with the desired expression system
    # ===========
    # negate values for reverse sorting
    targets_results = sorted( targets_results, key = lambda x: (-x['top_construct_data']['authenticity_score'], -x['nmatching_PDB_structures']) )

    # ===========
    # print and write file containing sorted targets with details on PDB structure selection
    # ===========

    print ''

    PDB_selections_text = ''
    PDB_selections_text += '= Targets sorted by number of PDB structures with matching expression system tag =\n\n'
    PDB_selections_text += '%18s  %24s  %16s  %23s  %18s  %30s  %11s  %12s\n' % ('target', 'nmatching_PDB_structures', 'top_PDB_chain_ID', 'detected_expression_tag', 'authenticity_score', 'expression_system', 'target_rank', 'target_score')
    PDB_selections_text += '%18s  %24s  %16s  %23s  %18s  %30s  %11s  %12s\n' % ('______', '________________________', '________________', '_______________________', '__________________', '_________________', '___________', '______________')

    ntargets_zero_or_more_PDB = 0
    ntargets_one_or_more_PDB = 0
    ntargets_zero_or_more_PDB_expr_tag = 0
    ntargets_one_or_more_PDB_expr_tag = 0
    for t, target_dict in enumerate(targets_results):
        # Add a line after the 96th target
        if t == 96:
            PDB_selections_text += ('=' * len(PDB_selections_text.split('\n')[2])) + '\n'

        # ignore targets with no matching PDB structures or plasmid sequence
        if target_dict['nmatching_PDB_structures'] == 0:
            continue

        targetID = target_dict['targetID']
        top_PDB_chain_ID = target_dict['top_PDB_chain_ID']
        top_construct_data = target_dict['top_construct_data']
        if top_construct_data['tag_type'] == None:
            expr_tag_string = 'None'
        else:
            expr_tag_string = top_construct_data['tag_type'] + '_' + top_construct_data['tag_loc']

        target_score = target_dict['DB_target_score']
        target_rank = target_dict['DB_target_rank']

        PDB_selections_text += '%18s  %24d  %16s  %23s  %18d  %30s  %11s  %12s\n' % (targetID, target_dict['nmatching_PDB_structures'], top_PDB_chain_ID, expr_tag_string, top_construct_data['authenticity_score'], top_construct_data['expression_system'], target_rank, target_score)
        if target_dict['nmatching_PDB_structures'] > 0:
            ntargets_zero_or_more_PDB += 1
            if top_construct_data['tag_type'] != None:
                ntargets_zero_or_more_PDB_expr_tag += 1
        if target_dict['nmatching_PDB_structures'] > 1:
            ntargets_one_or_more_PDB += 1
            if top_construct_data['tag_type'] != None:
                ntargets_one_or_more_PDB_expr_tag += 1

    PDB_selections_text += '\nTotal targets with plasmid and > 0 matching PDB structures: %d\n' % ntargets_zero_or_more_PDB
    PDB_selections_text += 'Total targets with plasmid and > 1 matching PDB structures: %d\n' % ntargets_one_or_more_PDB
    PDB_selections_text += 'Total targets with plasmid and > 0 matching PDB structures and a detected expr_tag: %d\n' % ntargets_zero_or_more_PDB_expr_tag
    PDB_selections_text += 'Total targets with plasmid and > 1 matching PDB structures and a detected expr_tag: %d\n' % ntargets_one_or_more_PDB_expr_tag

    print PDB_selections_text

    with open(output_selections_filepath, 'w') as output_selections_file:
        output_selections_file.write(PDB_selections_text)

    # ===========
    # write .xlsx spreadsheet containing the top 96 targets and necessary details for expression testing
    # ===========

    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = output_Excel_filename[0 : output_Excel_filename.find('.xlsx')]

    headings = ['targetID', 'GeneID', 'expression tag location', 'aa_start', 'aa_end', 'aa_seq', 'dna_start', 'dna_end', 'dna_seq']
    for h in range(len(headings)):
        heading_cell = ws.cell(row=0, column=h)
        heading_cell.value = headings[h]

    t_iter = 0
    ntargets_selected = 0

    # create new list of target results, removing targets with no matching PDB structures, and then keeping only the top [ndesired_unique_targets]
    xl_output_targets_results = [t for t in targets_results if t['nmatching_PDB_structures'] > 0][0:ndesired_unique_targets]

    # and a list of the same target results, sorted by DB rank
    xl_output_targets_results_sorted_by_target_rank = sorted( xl_output_targets_results, key = lambda x: int(x['DB_target_rank']) )

    for well_index in range(plate_size):
        if well_index < ndesired_unique_targets:
            target_index = well_index
            target_dict = xl_output_targets_results[target_index]
        else:
            target_index = well_index - ndesired_unique_targets
            target_dict = xl_output_targets_results_sorted_by_target_rank[target_index]

        targetID = target_dict['targetID']

        top_PDB_chain_ID = target_dict['top_PDB_chain_ID']

        target_NCBI_GeneID = target_dict['target_NCBI_GeneID']
        construct_aa_start = target_dict['construct_target_region_start_plasmid_coords']
        construct_aa_end = target_dict['construct_target_region_end_plasmid_coords'] # this refers to the last aa in 0-based aa coordinates
        construct_aa_seq = target_dict['construct_target_region_plasmid_seq']

        construct_dna_start = (construct_aa_start * 3)
        construct_dna_end = (construct_aa_end * 3) + 2 # this refers to the last nucleotide in 0-based nucleotide coordinates

        # Build spreadsheeet

        ID = targetID + '_' + top_PDB_chain_ID
        ID_cell = ws.cell(row=well_index+1, column=0)
        ID_cell.value = ID

        GeneID_cell = ws.cell(row=well_index+1, column=1)
        GeneID_cell.value = target_NCBI_GeneID
        
        exp_tag_loc = target_dict['top_construct_data']['tag_loc']
        exp_tag_loc_cell = ws.cell(row=well_index+1, column=2)
        exp_tag_loc_cell.value = exp_tag_loc

        # residue span: 1-based inclusive aa coordinates
        construct_aa_start_cell = ws.cell(row=well_index+1, column=3)
        construct_aa_end_cell = ws.cell(row=well_index+1, column=4)
        construct_aa_start_cell.value = construct_aa_start + 1
        construct_aa_end_cell.value = construct_aa_end + 1
        
        # aa_seq
        construct_aa_seq_cell = ws.cell(row=well_index+1, column=5)
        construct_aa_seq_cell.value = construct_aa_seq

        # DNA span: 1-based inclusive dna nucleotide coordinates
        construct_dna_start_cell = ws.cell(row=well_index+1, column=6)
        construct_dna_end_cell = ws.cell(row=well_index+1, column=7)
        construct_dna_start_cell.value = construct_dna_start + 1
        construct_dna_end_cell.value = construct_dna_end + 1
        
        # dna_seq
        orig_plasmid_dna_seq = plasmid_dna_seqs[target_NCBI_GeneID]
        construct_dna_seq = orig_plasmid_dna_seq[ construct_dna_start : construct_dna_end + 1 ]
        construct_dna_seq_cell = ws.cell(row=well_index+1, column=8)
        construct_dna_seq_cell.value = construct_dna_seq

        if len(construct_dna_seq) % 3 != 0:
            raise Exception, 'modulo 3 of DNA sequence length should be 0. Instead was %d' % (len(construct_dna_seq) % 3)

    # Save spreadsheet
    wb.save(output_Excel_filepath)

