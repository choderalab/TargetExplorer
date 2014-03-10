import os, re
import openpyxl
from openpyxl import Workbook
import Bio, Bio.Seq, Bio.Alphabet
from lxml import etree
from lxml.builder import E
import TargetExplorer as clab
import pandas as pd

plasmid_library_dir = '.'
Harvard_plasmid_library_filepath = os.path.join(plasmid_library_dir, 'Mehle_Kinase_VS1_pJP1520_new_plates.xlsx')

# ==========
# Parse plasmid library spreadsheet
# ==========

wb = openpyxl.load_workbook(Harvard_plasmid_library_filepath)
sheet_ranges = wb.get_sheet_by_name(name = 'Kinase_VS1_pJP1520_new_plates')
nrows = sheet_ranges.get_highest_row()
plasmidIDs = {}
plasmid_aa_seqs = {}
plasmid_dna_seqs = {}
for row in range(2,nrows):
    plasmidID = sheet_ranges.cell('E%d' % row).value    # type(int)
    NCBI_GeneID = sheet_ranges.cell('H%d' % row).value    # type(int)
    Symbol = sheet_ranges.cell('I%d' % row).value
    dna_seq = sheet_ranges.cell('L%d' % row).value 
    if len(dna_seq) % 3 != 0:
        print 'WARNING: length of DNA sequence not divisible by 3, for plasmid with Gene Symbol %s and NCBI Gene ID %s' % (Symbol, NCBI_GeneID)
    # Translate to DNA sequence (don't include stop codon)
    aa_seq = Bio.Seq.Seq(dna_seq, Bio.Alphabet.generic_dna).translate(to_stop=True)
    plasmidIDs[NCBI_GeneID] = plasmidID
    plasmid_aa_seqs[NCBI_GeneID] = aa_seq
    plasmid_dna_seqs[NCBI_GeneID] = dna_seq

plasmid_NCBI_GeneIDs = plasmid_aa_seqs.keys()

DB_path = os.path.join('..', '..', '..', '..', 'database', 'database.xml')

DB_root = etree.parse(DB_path).getroot()


# ===========
# Generate html header
# ===========

title = 'plasmids aligned against UniProt sequences'
subtitle = 'UniProt seqences are identified by targetID; plasmid sequences by GeneID. Second number on the plasmid row is the alignment score (plasmid seq vs UniProt seq) relative to the maximum possible (UniProt seq vs UniProt seq), ignoring gaps. Third number is the same plus a gap penalty of -8'

output_html_tree = E.html( E.head( E.link() ), E.body( E.h2(title), E.h3(subtitle), E.table() ) )
output_html_body = output_html_tree.find('body')
output_html_table = output_html_body.find('table')
css_filename = 'seqlib.css'
css_link = output_html_tree.find('head/link')
css_link.set('type','text/css')
css_path = os.path.join(css_filename)
css_link.set('href',css_path)
css_link.set('rel','stylesheet')

# ===========
# definitions
# ===========

def gen_html(html_table, aln_data):
    for i in range(len(aln_data)):
        row = E.tr()

        # row title
        row.append( E.td( E.div(aln_data[i][0], CLASS='ali') ) )


        # ===========
        # score the alignment quality, subtract from the maximum possible, and add to first row
        # ===========
        if i == 1:
            aln_qual_ignore_gaps = clab.align.score_aln(aln_data[0][1], aln_data[1][1], gap_penalty=0)
            aln_qual = clab.align.score_aln(aln_data[0][1], aln_data[1][1], gap_penalty=-8)
            max_aln_qual = clab.align.score_aln(aln_data[0][1], aln_data[0][1], gap_penalty=0)
            aln_score_ignore_gaps = str(aln_qual_ignore_gaps - max_aln_qual)
            aln_score = str(aln_qual - max_aln_qual)
        else:
            aln_score_ignore_gaps = ''
            aln_score = ''
        row.append( E.td( E.div(aln_score_ignore_gaps, CLASS='ali') ) )
        row.append( E.td( E.div(aln_score, CLASS='ali') ) )


        # format sequence with css classes. Returned as a list of span objects
        prettyseq = clab.core.seq2pretty_html(alignment[i][1])

        # set up sequence div
        seq_div = E.div(id='sequence',CLASS='ali')
        seq_div.set('style','background-color:#dddddd;letter-spacing:-5px')

        # add sequence to div
        for span in prettyseq:
            seq_div.append(span)

        row.append( E.td( seq_div, nowrap='' ) )

        html_table.append(row)

    html_table.append(E.tr())


# ===========
# Iterate through targets in DB
# ===========

DB_targets = DB_root.findall('entry/UniProt/domains/domain[@targetID]')

# To be used to construct a pandas DataFrame
output_data = {
'targetID':[],
'plasmidID':[],
'UniProt_canonseq_aln':[],
'plasmid_seq_aln':[],
'nconflicts':[],
'len_seq_aln':[],
'pctidentity':[]
}

for DB_domain in DB_targets:
    targetID = DB_domain.get('targetID')

    target_domain_len = int(DB_domain.get('length'))
    if target_domain_len > 350 or target_domain_len < 191:
        print 'WARNING: Target %s domain length %d.' % (targetID, target_domain_len)
    DB_entry = DB_domain.getparent().getparent().getparent()

    # get IDs
    target_UniProt_entry_name = DB_entry.find('UniProt').get('entry_name')
    target_NCBI_Gene_node = DB_entry.find('NCBI_Gene/entry[@ID]')
    if target_NCBI_Gene_node == None:
        print 'Gene ID not found for target %s' % target_UniProt_entry_name
        continue
    target_NCBI_GeneID = int(target_NCBI_Gene_node.get('ID'))

    if target_NCBI_GeneID not in plasmid_NCBI_GeneIDs:
        print 'Gene ID %s (%s) not found in Harvard plasmid library.' % (target_NCBI_GeneID, target_UniProt_entry_name)
        continue

    # get UniProt canonical isoform sequence
    UniProt_canonseq = clab.core.sequnwrap( DB_entry.findtext('UniProt/isoforms/canonical_isoform/sequence') )

    # =====
    # Run alignment
    # =====

    alnIDs = [targetID, str(target_NCBI_GeneID)]
    pre_aln_seqs = [UniProt_canonseq, plasmid_aa_seqs[target_NCBI_GeneID]]

    aligned_seqs = clab.align.run_clustalo(alnIDs, pre_aln_seqs)
    alignment = [ [alnIDs[i], seq] for i, seq in enumerate(aligned_seqs) ]

    # ===========
    # compare plasmid and PDB seqs with aligned UniProt canonical seq - put non-matching aas in lower case
    # ===========
    seq = list(alignment[1][1]) # sequence needs to be mutable so convert to list
    for aa_iter in range(len(seq)):
        if seq[aa_iter] != alignment[0][1][aa_iter]:
            seq[aa_iter] = seq[aa_iter].lower()
    alignment[1][1] = ''.join(seq)

    # ===========
    # generate html
    # ===========
    gen_html(output_html_table, alignment)

    # ===========
    # calculate number of conflicts and percentage
    # ===========
    nconflicts = len( [ aa_iter for aa_iter in range(len(alignment[0][1])) if alignment[0][1][aa_iter] == '-' or alignment[1][1][aa_iter] == '-' or re.match('[a-z]', alignment[1][1][aa_iter]) ] )
    pctidentity = (len(UniProt_canonseq) - nconflicts) * 100. / float(len(UniProt_canonseq))

    # ===========
    # append to data dict, to be used later to construct pandas DataFrame
    # ===========
    output_data['targetID'].append(targetID)
    plasmidID = plasmidIDs[target_NCBI_GeneID]
    output_data['plasmidID'].append(plasmidID)
    output_data['UniProt_canonseq_aln'].append(alignment[0][1])
    output_data['plasmid_seq_aln'].append(alignment[1][1])
    output_data['nconflicts'].append(str(nconflicts))
    output_data['len_seq_aln'].append(str(len(UniProt_canonseq)))
    output_data['pctidentity'].append('%.2f' % pctidentity)

    #break

# construct pandas DataFrame and write to csv
#df = pd.DataFrame(output_data)
#df.to_csv('aln-against-UniProt-seq.csv', index=False)

# write data txt file
with open('aln.txt', 'w') as otxtfile:
    columnwidths = {}
    for key in output_data.keys():
        columnwidths[key] = max( [ len(x) for x in output_data[key] ] )

    IDcolumnwidth = max( [columnwidths['targetID'], columnwidths['plasmidID']] )

    for i in range(len(output_data['targetID'])):
        otxtfile.write('%-*s' % (IDcolumnwidth, output_data['targetID'][i]))
        otxtfile.write('  ')
        otxtfile.write('%*s' % (columnwidths['pctidentity'], '') )
        otxtfile.write('  ')
        otxtfile.write('%*s' % (columnwidths['nconflicts'], '') )
        otxtfile.write(' ')
        otxtfile.write('%-*s' % (columnwidths['len_seq_aln'], '') )
        otxtfile.write('  ')
        otxtfile.write('%s' % output_data['UniProt_canonseq_aln'][i])
        otxtfile.write('\n')

        otxtfile.write('%-*s' % (IDcolumnwidth, output_data['plasmidID'][i]))
        otxtfile.write('  ')
        otxtfile.write('%*s' % (columnwidths['pctidentity'], output_data['pctidentity'][i]) )
        otxtfile.write('  ')
        otxtfile.write('%*s' % (columnwidths['nconflicts'], output_data['nconflicts'][i]) )
        otxtfile.write('/')
        otxtfile.write('%-*s' % (columnwidths['len_seq_aln'], output_data['len_seq_aln'][i]) )
        otxtfile.write('  ')
        otxtfile.write('%s' % output_data['plasmid_seq_aln'][i])
        otxtfile.write('\n\n')

with open('aln.html', 'w') as ohtmlfile:
    ohtmlfile.write( etree.tostring(output_html_tree, pretty_print=True) )

